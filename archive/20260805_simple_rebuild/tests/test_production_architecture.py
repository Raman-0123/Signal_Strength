import io
import tempfile
import unittest
from concurrent.futures import ThreadPoolExecutor
from dataclasses import replace
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from core.google_search import GoogleSecurityCheck
from speedy_scraper.config import load_catalog, load_config
from speedy_scraper.dedup import parse_dedup_content
from speedy_scraper.domain import (
    CandidateEvaluation,
    DedupKey,
    DedupScope,
    JobStatus,
    LeadRecord,
    ProxyConfig,
    RateLimitConfig,
    ResultKind,
    RetryConfig,
    SearchPage,
    SearchResult,
    StorageConfig,
)
from speedy_scraper.events import redact
from speedy_scraper.exports import LEAD_EXPORT_COLUMNS, export_artifacts, export_leads
from speedy_scraper.orchestrator import ScraperOrchestrator
from speedy_scraper.reliability import (
    ProviderBlockedError,
    ProxyPool,
    RetryExecutor,
    TokenBucketRateLimiter,
    parse_retry_after,
)
from speedy_scraper.repository import LeadRepository


class _FakeProvider:
    name = "fake"

    def __init__(self):
        self.requests = []

    def search(self, request):
        self.requests.append(request)
        return SearchPage(
            results=(SearchResult(
                "Jane Doe - CMO at Acme | LinkedIn",
                "Location: Delhi, India",
                "https://www.linkedin.com/in/jane-doe/",
            ),),
            provider="fake",
            page=request.page,
        )


class _ValidationCaptchaProvider(_FakeProvider):
    def __init__(self):
        super().__init__()
        self.challenged = False

    def search(self, request):
        if request.linkedin_only:
            return super().search(request)
        self.requests.append(request)
        if not self.challenged:
            self.challenged = True
            raise GoogleSecurityCheck(
                request.query,
                "https://www.google.com/sorry/index",
                page=request.page,
                linkedin_only=False,
                max_results=request.max_results,
            )
        return SearchPage(
            results=(SearchResult(
                "Acme software company",
                "Acme develops enterprise software and technology platforms.",
                "https://acme.example/about",
            ),),
            provider="fake",
            page=request.page,
        )


class _DiscoveryBlockedProvider(_FakeProvider):
    def search(self, request):
        self.requests.append(request)
        raise ProviderBlockedError("DuckDuckGo 429 challenge", provider="ddgs", retry_after=0)


class _ValidationBlockedProvider(_FakeProvider):
    def __init__(self):
        super().__init__()
        self.blocked = False

    def search(self, request):
        if request.linkedin_only:
            return super().search(request)
        self.requests.append(request)
        if not self.blocked:
            self.blocked = True
            raise ProviderBlockedError("DuckDuckGo 429 challenge", provider="ddgs", retry_after=0)
        return SearchPage(
            results=(SearchResult(
                "Acme software company",
                "Acme develops enterprise software and technology platforms.",
                "https://acme.example/about",
            ),),
            provider="fake",
            page=request.page,
        )


class ProductionArchitectureTests(unittest.TestCase):
    def test_configuration_precedence_and_secret_policy(self):
        with tempfile.TemporaryDirectory() as directory:
            settings = Path(directory) / "settings.yaml"
            settings.write_text(
                "storage:\n  database_path: user.db\nretry:\n  max_attempts: 4\n",
                encoding="utf-8",
            )
            config = load_config(
                settings,
                env={
                    "SPEEDY_SCRAPER_DB_PATH": "environment.db",
                    "SPEEDY_SCRAPER_RETRY__MAX_ATTEMPTS": "5",
                },
                overrides={"storage": {"database_path": "request.db"}},
            )
            self.assertEqual(config.storage.database_path, "request.db")
            self.assertEqual(config.retry.max_attempts, 5)

            settings.write_text("api_key: do-not-store-this\n", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "environment variables"):
                load_config(settings, env={})

    def test_configuration_validation_environment_and_catalog_edges(self):
        invalid_overrides = [
            {"retry": {"max_attempts": 0}},
            {"storage": {"retention_days": 0}},
            {"query_budget": {"acceptance_rate": 0}},
            {"query_budget": {"citations_per_query": 0}},
            {"query_budget": {"maximum_queries": 0}},
            {"scheduler": {"max_workers": 0}},
            {"rate_limits": {"ddgs": {"requests_per_minute": 0, "minimum_interval_seconds": 0}}},
            {"rate_limits": {"ddgs": {"requests_per_minute": 1, "minimum_interval_seconds": -1}}},
            {"browser": {"captcha_poll_seconds": -1}},
            {"browser": {"page_settle_min_seconds": 2, "page_settle_max_seconds": 1}},
            {"browser": {"post_search_min_seconds": 2, "post_search_max_seconds": 1}},
            {"proxy": {"failure_threshold": 0}},
            {"proxy": {"urls": ["ftp://unsupported"]}},
        ]
        for override in invalid_overrides:
            with self.subTest(override=override), self.assertRaises(ValueError):
                load_config(env={}, overrides=override)

        config = load_config(env={
            "SPEEDY_SCRAPER_DATA_DIR": "custom-data",
            "SPEEDY_SCRAPER_LOG_LEVEL": "debug",
            "SPEEDY_SCRAPER_API_HOST": "localhost",
            "SPEEDY_SCRAPER_API_PORT": "9000",
            "SPEEDY_SCRAPER_API_KEY": "secret",
            "SPEEDY_SCRAPER_CHROME_PATH": "/chrome",
            "SPEEDY_SCRAPER_RETENTION_DAYS": "12",
            "SPEEDY_SCRAPER_SCHEDULER_TIMEZONE": "UTC",
            "SPEEDY_SCRAPER_SCHEDULED_HEADLESS": "off",
            "SPEEDY_SCRAPER_PROXY_URLS": "http://a:1, socks5://b:2",
            "NOT_SPEEDY": "ignored",
        })
        self.assertEqual(config.api_port, 9000)
        self.assertFalse(config.browser.scheduled_headless)
        self.assertEqual(len(config.proxy.urls), 2)
        with self.assertRaisesRegex(ValueError, "boolean"):
            load_config(env={"SPEEDY_SCRAPER_SCHEDULED_HEADLESS": "maybe"})

        catalog = load_catalog()
        resolved = catalog.resolve("roles", ["CXO", "CXO", "Novel Role"])
        self.assertIn("Novel Role", resolved)
        self.assertEqual(len(resolved), len(set(resolved)))
        with tempfile.TemporaryDirectory() as directory:
            missing = Path(directory) / "missing.yaml"
            with self.assertRaisesRegex(ValueError, "not found"):
                load_config(missing, env={})
            invalid = Path(directory) / "invalid.yaml"
            invalid.write_text("key: [unterminated", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "Invalid YAML"):
                load_config(invalid, env={})
            root_list = Path(directory) / "list.yaml"
            root_list.write_text("- value\n", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "root must be a mapping"):
                load_config(root_list, env={})
            bad_catalog = Path(directory) / "catalog.yaml"
            bad_catalog.write_text("locations: {}\n", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "Catalog category"):
                load_catalog(bad_catalog)

    def test_retry_full_jitter_rate_limit_and_proxy_cooldown(self):
        sleeps = []
        attempts = []
        retry = RetryExecutor(
            RetryConfig(max_attempts=3, base_delay_seconds=1, jitter_ratio=1),
            sleeper=sleeps.append,
            randomizer=lambda: 0.25,
        )

        def flaky():
            attempts.append(1)
            if len(attempts) < 3:
                raise TimeoutError("temporary")
            return "ok"

        value, count = retry.run(flaky, provider="test")
        self.assertEqual((value, count), ("ok", 3))
        self.assertEqual(sleeps, [0.25, 0.5])

        now = [0.0]
        limiter_sleeps = []

        def sleep_and_advance(delay):
            limiter_sleeps.append(delay)
            now[0] += delay

        limiter = TokenBucketRateLimiter(
            {"test": RateLimitConfig(60, 2)},
            clock=lambda: now[0], sleeper=sleep_and_advance,
        )
        limiter.acquire("test")
        limiter.acquire("test")
        self.assertEqual(limiter_sleeps, [2.0])

        proxy = "http://person:secret@proxy.example:8080"
        pool = ProxyPool(
            ProxyConfig(
                enabled=True, urls=(proxy, "socks5://backup.example:1080"),
                failure_threshold=1, cooldown_seconds=30,
            ),
            clock=lambda: now[0],
        )
        first = pool.acquire("job-one")
        pool.report_failure("job-one", rotate=True)
        self.assertNotEqual(pool.acquire("job-one"), first)
        self.assertNotIn("secret", str(redact({"proxy": proxy})))
        now_http = datetime(2026, 8, 4, 12, 0, tzinfo=timezone.utc)
        self.assertEqual(parse_retry_after("12", now=now_http), 12.0)
        self.assertEqual(
            parse_retry_after("Tue, 04 Aug 2026 12:00:15 GMT", now=now_http),
            15.0,
        )
        self.assertIsNone(parse_retry_after("not-a-date", now=now_http))

    def test_repository_state_machine_atomic_checkpoint_and_global_dedup(self):
        with tempfile.TemporaryDirectory() as directory:
            repository = LeadRepository(Path(directory) / "speedy.db")
            repository.migrate()
            request = {"workflow": "lead", "query_plan": [{"query": "q"}]}
            job = repository.create_job("lead", request)
            repository.transition(job.id, JobStatus.RUNNING)
            lead = LeadRecord(
                name="Jane Doe", designation="CMO", company="Acme",
                linkedin_url="https://linkedin.com/in/jane-doe",
                verified_location="Delhi",
            )
            checkpoint = {
                "query_index": 0,
                "query_state": {"result_index": 0},
                "browser_state": {},
            }
            lead_id = repository.save_qualified(
                job.id,
                lead,
                checkpoint=checkpoint,
                pending_validation={"candidate": {"name": "Jane Doe"}},
            )
            persisted = repository.get_job(job.id)
            self.assertEqual(persisted.checkpoint["pending_validation"]["lead_id"], lead_id)
            self.assertTrue(repository.contains_dedup_key("url", lead.linkedin_url))
            self.assertTrue(repository.contains_dedup_key(
                DedupKey("url", lead.linkedin_url), DedupScope.JOB, job_id=job.id,
            ))
            with self.assertRaisesRegex(ValueError, "Invalid job transition"):
                repository.transition(job.id, JobStatus.QUEUED)

    def test_atomic_lane_claims_stale_lease_recovery_and_retention(self):
        with tempfile.TemporaryDirectory() as directory:
            repository = LeadRepository(StorageConfig(
                database_path=str(Path(directory) / "leases.db"), retention_days=0,
            ))
            repository.migrate()
            first = repository.create_job("reconcile", {"workflow": "reconcile"})
            repository.create_job("reconcile", {"workflow": "reconcile"})
            google = repository.create_job("lead", {
                "workflow": "lead", "discovery_provider": "google",
                "validation_provider": "ddgs",
            })

            with ThreadPoolExecutor(max_workers=2) as pool:
                claims = list(pool.map(
                    lambda runner: repository.claim_next_job(runner, lane="non_google"),
                    ("runner-a", "runner-b"),
                ))
            claimed_ids = {item.id for item in claims if item is not None}
            self.assertEqual(len(claimed_ids), 2)
            google_claim = repository.claim_next_job("google-runner", lane="google")
            self.assertEqual(google_claim.id, google.id)

            with repository.transaction(immediate=True) as connection:
                connection.execute(
                    "UPDATE jobs SET lease_expires_at=? WHERE id=?",
                    ("2000-01-01T00:00:00+00:00", first.id),
                )
            self.assertGreaterEqual(repository.recover_stale_jobs(), 1)
            self.assertEqual(repository.get_job(first.id).status, JobStatus.QUEUED)

            retained = repository.get_job(next(iter(claimed_ids - {first.id})))
            lead = LeadRecord(
                name="Persistent Lead", designation="CEO", company="Durable Co",
                linkedin_url="https://linkedin.com/in/persistent-lead",
                verified_location="Mumbai",
            )
            repository.save_qualified(retained.id, lead)
            repository.add_event(retained.id, "info", "detail", "old detail")
            repository.increment_metric(retained.id, "queries", 1)
            repository.save_evidence(retained.id, "company", "durable co", "ddgs", "q", "e", 1)
            repository.save_checkpoint(retained.id, {"query_index": 0, "query_state": {}, "browser_state": {}})
            repository.transition(retained.id, JobStatus.COMPLETED)
            cleanup = repository.cleanup()
            self.assertGreaterEqual(cleanup["events"], 1)
            self.assertEqual(repository.list_events(retained.id), [])
            self.assertEqual(repository.metrics(retained.id), {})
            self.assertEqual(repository.get_job(retained.id).checkpoint, {})
            self.assertTrue(repository.contains_dedup_key("url", lead.linkedin_url))

    def test_new_lead_jobs_are_ddgs_only(self):
        with tempfile.TemporaryDirectory() as directory:
            base = load_config(env={})
            config = replace(
                base,
                storage=StorageConfig(database_path=str(Path(directory) / "ddgs-only.db")),
            )
            repository = LeadRepository(config.storage)
            repository.migrate()
            orchestrator = ScraperOrchestrator(config, repository, catalog=load_catalog())
            request = {
                "workflow": "lead",
                "locations": ["Delhi"],
                "roles": ["CMO"],
                "target_count": 1,
            }
            job = orchestrator.create_job({**request, "search_provider": "ddgs"})
            self.assertEqual(job.request["discovery_provider"], "ddgs")
            self.assertEqual(job.request["validation_provider"], "ddgs")
            for field in ("search_provider", "discovery_provider", "validation_provider"):
                with self.subTest(field=field), self.assertRaisesRegex(ValueError, "only support ddgs"):
                    orchestrator.create_job({**request, field: "google"})
                with self.subTest(field=f"{field}-brave"), self.assertRaisesRegex(ValueError, "only support ddgs"):
                    orchestrator.create_job({**request, field: "brave"})

    def test_export_contracts_keep_four_lead_fields_and_csev_sheet(self):
        with tempfile.TemporaryDirectory() as directory:
            repository = LeadRepository(Path(directory) / "exports.db")
            repository.migrate()
            lead_job = repository.create_job("lead", {"workflow": "lead"})
            lead = LeadRecord(
                name="Jane Doe", designation="CMO", company="Acme",
                linkedin_url="https://linkedin.com/in/jane-doe",
                verified_location="Delhi",
            )
            lead_id = repository.save_qualified(lead_job.id, lead)
            repository.update_strict_result(
                lead_job.id, lead_id,
                CandidateEvaluation(hard_qualified=True, strict_qualified=True),
            )
            csv_bytes, media_type, _name = export_leads(
                repository, lead_job.id, "csv", ResultKind.QUALIFIED,
            )
            self.assertEqual(media_type, "text/csv")
            self.assertEqual(csv_bytes.decode().splitlines()[0].split(","), LEAD_EXPORT_COLUMNS)
            xlsx_bytes, _media_type, _name = export_leads(
                repository, lead_job.id, "xlsx", ResultKind.QUALIFIED,
            )
            workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
            self.assertEqual(workbook.sheetnames, ["All Qualified POCs", "Strict Matches"])
            self.assertEqual(
                [cell.value for cell in next(workbook["All Qualified POCs"].iter_rows())],
                LEAD_EXPORT_COLUMNS,
            )

            competitor_job = repository.create_job("competitor", {"workflow": "competitor"})
            repository.save_artifact(competitor_job.id, "competitor_leads", [{
                "Full_Name": "Alex", "Company": "Acme", "Designation": "VP",
                "LinkedIn_URL": "https://linkedin.com/in/alex",
            }])
            repository.save_artifact(competitor_job.id, "company_summaries", {
                "Acme": {"description": "Software", "net_profit": "Unknown"},
            })
            csev_bytes, _media_type, _name = export_artifacts(repository, competitor_job.id, "xlsx")
            csev = load_workbook(io.BytesIO(csev_bytes), read_only=True)
            self.assertEqual(csev.sheetnames, ["CSEV"])

    def test_engine_has_no_streamlit_state_and_stops_on_qualified_target(self):
        with tempfile.TemporaryDirectory() as directory:
            base = load_config(env={})
            config = replace(
                base,
                storage=StorageConfig(database_path=str(Path(directory) / "engine.db")),
            )
            repository = LeadRepository(config.storage)
            repository.migrate()
            orchestrator = ScraperOrchestrator(
                config, repository, catalog=load_catalog(),
            )
            job = orchestrator.create_job({
                "workflow": "lead",
                "locations": ["Delhi"],
                "roles": ["CMO"],
                "target_count": 1,
                "discovery_provider": "ddgs",
                "validation_provider": "ddgs",
                "edited_queries": ['site:linkedin.com/in "CMO" "Delhi"'],
            })
            provider = _FakeProvider()
            orchestrator.engine._provider = lambda *_args: provider
            result = orchestrator.engine.run(job.id)
            self.assertEqual(result.job.status, JobStatus.COMPLETED)
            self.assertEqual(len(result.qualified), 1)
            self.assertEqual(
                result.qualified[0].as_export_row(),
                {
                    "Name": "Jane Doe", "Designation": "CMO",
                    "Company": "Acme", "Location": "Delhi, India",
                },
            )

    def test_excel_all_sheets_are_imported_for_exclusion(self):
        content = io.BytesIO()
        with pd.ExcelWriter(content, engine="openpyxl") as writer:
            pd.DataFrame({
                "Name": ["Jane Doe"], "Company": ["Acme"],
                "LinkedIn": ["https://linkedin.com/in/jane-doe"],
            }).to_excel(writer, index=False, sheet_name="Prior")
            pd.DataFrame({
                "Name": ["Bob Singh"], "Company": ["Beta"],
                "LinkedIn": ["https://linkedin.com/in/bob-singh"],
            }).to_excel(writer, index=False, sheet_name="CRM")
        keys, sheet_count = parse_dedup_content("history.xlsx", content.getvalue())
        self.assertEqual(sheet_count, 2)
        self.assertEqual(len([key for key in keys if key[0] == "url"]), 2)

    def test_discovery_ddgs_block_checkpoints_without_advancing(self):
        with tempfile.TemporaryDirectory() as directory:
            base = load_config(env={})
            config = replace(
                base,
                storage=StorageConfig(database_path=str(Path(directory) / "discovery-block.db")),
            )
            repository = LeadRepository(config.storage)
            repository.migrate()
            orchestrator = ScraperOrchestrator(config, repository, catalog=load_catalog())
            job = orchestrator.create_job({
                "workflow": "lead",
                "locations": ["Delhi"],
                "roles": ["CMO"],
                "target_count": 1,
                "edited_queries": ['site:linkedin.com/in "CMO" "Delhi"'],
            })
            provider = _DiscoveryBlockedProvider()
            orchestrator.engine._provider = lambda *_args: provider
            waiting = orchestrator.engine.step(job.id).job
            self.assertEqual(waiting.status, JobStatus.WAITING_VERIFICATION)
            self.assertEqual(waiting.checkpoint["query_index"], 0)
            self.assertEqual(waiting.checkpoint["query_state"]["result_index"], 0)
            self.assertEqual(waiting.checkpoint["security_check"]["engine"], "DuckDuckGo")
            self.assertEqual(repository.metrics(job.id)["ddgs_block_events"], 1)
            resumed = orchestrator.poll_verification(job.id)
            self.assertEqual(resumed.status, JobStatus.QUEUED)
            self.assertIsNone(resumed.checkpoint["security_check"])

    def test_ddgs_waiting_manual_resume_retry_and_future_cooldown(self):
        with tempfile.TemporaryDirectory() as directory:
            base = load_config(env={})
            config = replace(
                base,
                storage=StorageConfig(database_path=str(Path(directory) / "manual-ddgs.db")),
            )
            repository = LeadRepository(config.storage)
            repository.migrate()
            orchestrator = ScraperOrchestrator(config, repository, catalog=load_catalog())
            checkpoint = {
                "query_index": 0,
                "query_state": {"page": 1},
                "browser_state": {},
                "security_check": {
                    "engine": "DuckDuckGo",
                    "query": "saved query",
                    "retry_at": "2999-01-01T00:00:00+00:00",
                },
            }
            waiting = repository.create_job("lead", {
                "workflow": "lead",
                "discovery_provider": "ddgs",
                "validation_provider": "ddgs",
            }, checkpoint=checkpoint)
            repository.transition(waiting.id, JobStatus.RUNNING)
            repository.transition(waiting.id, JobStatus.WAITING_VERIFICATION)
            self.assertEqual(orchestrator.poll_verification(waiting.id).status, JobStatus.WAITING_VERIFICATION)
            resumed = orchestrator.resume_job(waiting.id)
            self.assertEqual(resumed.status, JobStatus.QUEUED)
            self.assertIsNone(resumed.checkpoint["security_check"])

            retry_waiting = repository.create_job("lead", {
                "workflow": "lead",
                "discovery_provider": "ddgs",
                "validation_provider": "ddgs",
            }, checkpoint=checkpoint)
            repository.transition(retry_waiting.id, JobStatus.RUNNING)
            repository.transition(retry_waiting.id, JobStatus.WAITING_VERIFICATION)
            retried = orchestrator.retry_job(retry_waiting.id)
            self.assertEqual(retried.status, JobStatus.QUEUED)
            self.assertIsNone(retried.checkpoint["security_check"])

    def test_validation_ddgs_block_resumes_pending_candidate_without_page_replay(self):
        with tempfile.TemporaryDirectory() as directory:
            base = load_config(env={})
            config = replace(
                base,
                storage=StorageConfig(database_path=str(Path(directory) / "blocked.db")),
            )
            repository = LeadRepository(config.storage)
            repository.migrate()
            orchestrator = ScraperOrchestrator(config, repository, catalog=load_catalog())
            job = orchestrator.create_job({
                "workflow": "lead",
                "locations": ["Delhi"],
                "roles": ["CMO"],
                "industries": ["Software"],
                "target_count": 1,
                "discovery_provider": "ddgs",
                "validation_provider": "ddgs",
                "edited_queries": ['site:linkedin.com/in "CMO" "Delhi"'],
            })
            provider = _ValidationBlockedProvider()
            orchestrator.engine._provider = lambda *_args: provider
            orchestrator.engine.step(job.id)
            waiting = orchestrator.engine.step(job.id).job
            self.assertEqual(waiting.status, JobStatus.WAITING_VERIFICATION)
            self.assertEqual(waiting.qualified_count, 1)
            self.assertIsNotNone(waiting.checkpoint["pending_validation"])
            self.assertEqual(waiting.checkpoint["query_state"]["result_index"], 0)
            self.assertEqual(waiting.checkpoint["security_check"]["engine"], "DuckDuckGo")

            resumed = orchestrator.poll_verification(job.id)
            self.assertEqual(resumed.status, JobStatus.QUEUED)
            result = orchestrator.engine.run(job.id)
            self.assertEqual(result.job.status, JobStatus.COMPLETED)
            self.assertEqual(len(result.qualified), 1)
            self.assertEqual(len(result.strict), 1)
            discovery_requests = [item for item in provider.requests if item.linkedin_only]
            self.assertEqual(len(discovery_requests), 1)

    def test_lost_verification_browser_requeues_exact_checkpoint(self):
        with tempfile.TemporaryDirectory() as directory:
            base = load_config(env={})
            config = replace(
                base,
                storage=StorageConfig(database_path=str(Path(directory) / "lost.db")),
            )
            repository = LeadRepository(config.storage)
            repository.migrate()
            orchestrator = ScraperOrchestrator(config, repository, catalog=load_catalog())
            job = repository.create_job("lead", {
                "workflow": "lead", "discovery_provider": "google",
                "validation_provider": "google", "query_plan": [{"query": "exact query"}],
            }, checkpoint={
                "query_index": 0,
                "query_state": {"page": 3, "result_index": 7},
                "pending_validation": {"lead_id": "pending"},
                "browser_state": {"pid": 999999, "profile_dir": "/tmp/exact"},
                "security_check": {"query": "exact query", "page": 3},
            })
            repository.transition(job.id, JobStatus.RUNNING)
            repository.transition(job.id, JobStatus.WAITING_VERIFICATION)
            with (
                patch.object(orchestrator.browser_manager, "verification_resolved", return_value=False),
                patch.object(orchestrator.browser_manager, "is_running", return_value=False),
                patch.object(orchestrator.browser_manager, "display_available", return_value=True),
            ):
                resumed = orchestrator.poll_verification(job.id)
            self.assertEqual(resumed.status, JobStatus.QUEUED)
            self.assertEqual(resumed.checkpoint["query_index"], 0)
            self.assertEqual(resumed.checkpoint["query_state"], {"page": 3, "result_index": 7})
            self.assertEqual(resumed.checkpoint["pending_validation"], {"lead_id": "pending"})
            self.assertIsNone(resumed.checkpoint["security_check"])


if __name__ == "__main__":
    unittest.main()
