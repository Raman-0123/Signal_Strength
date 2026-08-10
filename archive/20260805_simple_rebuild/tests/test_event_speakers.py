import io
import json
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import Mock, patch

import requests
from openpyxl import load_workbook

from core.utils import normalize_linkedin_url
from speedy_scraper.config import load_config
from speedy_scraper.domain import JobStatus, SearchPage, SearchResult
from speedy_scraper.event_speakers import (
    EventSpeaker,
    EventSpeakerEngine,
    NoSpeakersFoundError,
    ResponseTooLargeError,
    SafeHtmlFetcher,
    UnsafeSourceUrlError,
    choose_speaker_match,
    extract_event_speakers,
    linkedin_id_from_url,
    validate_public_source_url,
)
from speedy_scraper.exports import EVENT_SPEAKER_EXPORT_COLUMNS, export_artifacts
from speedy_scraper.orchestrator import ScraperOrchestrator
from speedy_scraper.reliability import ProxyPool, RetryExecutor
from speedy_scraper.repository import LeadRepository


def _flight_html(rows, *, split_at=None):
    decoded = 'prefix,"allSpeakersData":' + json.dumps({"data": rows}) + ",suffix"
    chunks = [decoded]
    if split_at is not None:
        chunks = [decoded[:split_at], decoded[split_at:]]
    scripts = [f"<script>self.__next_f.push({json.dumps([1, chunk])})</script>" for chunk in chunks]
    return "<html><body>" + "".join(scripts) + "</body></html>"


def _row(
    name,
    *,
    speaker_id=None,
    company="Acme",
    designation="CEO",
    linkedin=None,
    active=True,
    country="India",
):
    return {
        "speakerId": speaker_id or name.lower().replace(" ", "-"),
        "fullName": name,
        "desgination": designation,
        "companyName": company,
        "linkedinProfile": linkedin,
        "isActive": active,
        "country": {"country": country},
    }


class _PublicResolver:
    def __call__(self, host, port, **_kwargs):
        address = "127.0.0.1" if host in {"localhost", "127.0.0.1", "private.example"} else "8.8.8.8"
        return [(2, 1, 6, "", (address, port))]


class _Limiter:
    def __init__(self):
        self.providers = []

    def acquire(self, provider):
        self.providers.append(provider)


def _response(status=200, content=b"<html></html>", content_type="text/html"):
    response = requests.Response()
    response.status_code = status
    response._content = content
    response._content_consumed = True
    response.headers["Content-Type"] = content_type
    response.encoding = "utf-8"
    return response


class EventSpeakerExtractionTests(unittest.TestCase):
    def test_extracts_split_flight_payload_filters_and_deduplicates(self):
        rows = [
            _row(
                "Jane  Doe",
                linkedin=(
                    "https://www.linkedin.com/mwlite/profile/in/Jane-Doe"
                    "?utm_source=share"
                ),
            ),
            _row("Jane Doe", speaker_id="duplicate", linkedin=None),
            _row("Inactive Person", active=False),
            _row("Bob Singh", company="Beta", linkedin="https://linkedin.com/company/beta"),
        ]
        html = _flight_html(rows, split_at=35)

        speakers = extract_event_speakers(html, "https://events.example/speakers")

        self.assertEqual([item.name for item in speakers], ["Jane Doe", "Bob Singh"])
        self.assertEqual(speakers[0].linkedin_url, "https://www.linkedin.com/in/jane-doe/")
        self.assertEqual(speakers[0].match_status, "provided")
        self.assertEqual(speakers[0].country, "India")
        self.assertEqual(speakers[1].linkedin_url, "")
        self.assertEqual(speakers[1].match_status, "not_found")
        self.assertEqual(speakers[1].match_evidence, "invalid provided LinkedIn URL")

    def test_missing_or_empty_payload_fails_explicitly(self):
        with self.assertRaises(NoSpeakersFoundError):
            extract_event_speakers("<html></html>", "https://example.com")
        with self.assertRaises(NoSpeakersFoundError):
            extract_event_speakers(_flight_html([]), "https://example.com")

    def test_linkedin_normalization_and_id_contract(self):
        legacy = "https://www.linkedin.com/mwlite/profile/in/Jane-Doe?trk=share"
        self.assertEqual(normalize_linkedin_url(legacy), "https://www.linkedin.com/in/jane-doe/")
        self.assertEqual(linkedin_id_from_url(legacy), "jane-doe")
        for invalid in (
            "https://linkedin.com/company/acme",
            "https://linkedin.com/posts/example",
            "https://example.com/in/jane",
        ):
            self.assertEqual(normalize_linkedin_url(invalid), "")
            self.assertEqual(linkedin_id_from_url(invalid), "")


class EventSpeakerSafetyTests(unittest.TestCase):
    def setUp(self):
        self.limiter = _Limiter()
        self.retry = RetryExecutor(load_config(env={}).retry, sleeper=lambda _delay: None)
        self.proxies = ProxyPool(load_config(env={}).proxy)

    def test_url_validation_rejects_unsafe_targets_and_credentials(self):
        resolver = _PublicResolver()
        self.assertEqual(
            validate_public_source_url("https://events.example/speakers", resolver=resolver),
            "https://events.example/speakers",
        )
        for url in (
            "file:///tmp/speakers",
            "http://localhost/speakers",
            "http://private.example/speakers",
            "https://user:secret@events.example/speakers",
        ):
            with self.assertRaises(UnsafeSourceUrlError, msg=url):
                validate_public_source_url(url, resolver=resolver)

    def test_fetcher_revalidates_redirect_and_bounds_response(self):
        redirect = _response(302)
        redirect.headers["Location"] = "http://private.example/speakers"
        session = Mock()
        session.headers = {}
        session.get.return_value = redirect
        fetcher = SafeHtmlFetcher(
            self.limiter,
            self.retry,
            self.proxies,
            session=session,
            resolver=_PublicResolver(),
        )
        with self.assertRaises(UnsafeSourceUrlError):
            fetcher.fetch("https://events.example/speakers", job_id="job")
        self.assertEqual(session.get.call_count, 1)

        oversized = _response(content=b"0123456789")
        oversized.headers["Content-Length"] = "10"
        session.get.return_value = oversized
        limited = SafeHtmlFetcher(
            self.limiter,
            self.retry,
            self.proxies,
            session=session,
            resolver=_PublicResolver(),
            max_bytes=5,
        )
        with self.assertRaises(ResponseTooLargeError):
            limited.fetch("https://events.example/speakers", job_id="job")

    def test_fetcher_requires_html(self):
        session = Mock()
        session.headers = {}
        session.get.return_value = _response(content=b"{}", content_type="application/json")
        fetcher = SafeHtmlFetcher(
            self.limiter,
            self.retry,
            self.proxies,
            session=session,
            resolver=_PublicResolver(),
        )
        with self.assertRaisesRegex(ValueError, "must return HTML"):
            fetcher.fetch("https://events.example/speakers", job_id="job")


class EventSpeakerMatchingTests(unittest.TestCase):
    def speaker(self, **kwargs):
        return EventSpeaker(
            speaker_id="speaker",
            name=kwargs.get("name", "Jane Doe"),
            designation=kwargs.get("designation", "CEO"),
            company=kwargs.get("company", "Acme"),
            country="India",
            source_url="https://events.example/speakers",
        )

    def test_exact_name_and_company_is_accepted(self):
        result = SearchResult(
            "Jane Doe - Chief Executive Officer at Acme | LinkedIn",
            "Jane Doe leads Acme.",
            "https://in.linkedin.com/in/jane-doe?trk=search",
        )
        matched = choose_speaker_match(self.speaker(), [result])
        self.assertEqual(matched.match_status, "matched")
        self.assertGreaterEqual(matched.confidence, 0.8)
        self.assertEqual(matched.linkedin_url, "https://www.linkedin.com/in/jane-doe/")

    def test_designation_fallback_can_be_accepted(self):
        result = SearchResult(
            "Jane Doe - CTO | LinkedIn",
            "Technology leader",
            "https://linkedin.com/in/jane-tech",
        )
        matched = choose_speaker_match(self.speaker(designation="CTO", company=""), [result])
        self.assertEqual(matched.match_status, "matched")
        self.assertEqual(matched.confidence, 0.8)

    def test_tied_candidates_are_ambiguous_and_no_result_is_not_found(self):
        results = [
            SearchResult(
                "Jane Doe - CEO at Acme | LinkedIn",
                "Acme",
                f"https://linkedin.com/in/jane-{suffix}",
            )
            for suffix in ("one", "two")
        ]
        ambiguous = choose_speaker_match(self.speaker(), results)
        self.assertEqual(ambiguous.match_status, "ambiguous")
        self.assertEqual(ambiguous.linkedin_url, "")
        self.assertIn("candidate=", ambiguous.match_evidence)

        missing = choose_speaker_match(self.speaker(), [
            SearchResult("Another Person - CEO", "Acme", "https://linkedin.com/in/another"),
        ])
        self.assertEqual(missing.match_status, "not_found")
        self.assertEqual(missing.linkedin_url, "")


class EventSpeakerWorkflowTests(unittest.TestCase):
    def _services(self, directory):
        config = load_config(overrides={
            "storage": {"database_path": str(Path(directory) / "speakers.db")},
            "rate_limits": {
                "event_page": {"requests_per_minute": 1000, "minimum_interval_seconds": 0},
                "ddgs": {"requests_per_minute": 1000, "minimum_interval_seconds": 0},
            },
        }, env={})
        repository = LeadRepository(config.storage)
        repository.migrate()
        orchestrator = ScraperOrchestrator(config, repository)
        return config, repository, orchestrator

    def test_persisted_workflow_results_metrics_and_exports(self):
        with tempfile.TemporaryDirectory() as directory:
            _config, repository, orchestrator = self._services(directory)
            source_url = "https://events.example/speakers"
            orchestrator.event_speaker_engine.fetcher = SimpleNamespace(
                fetch=lambda _url, job_id: _flight_html([
                    _row(
                        "Provided Person",
                        linkedin="https://linkedin.com/in/provided-person",
                    ),
                    _row("Jane Doe", linkedin=None),
                ])
            )
            provider = Mock()
            provider.search.return_value = SearchPage(
                results=(SearchResult(
                    "Jane Doe - CEO at Acme | LinkedIn",
                    "Jane Doe is CEO at Acme",
                    "https://linkedin.com/in/jane-doe",
                ),),
                provider="ddgs",
                page=1,
            )
            registry = Mock()
            registry.get.return_value = provider
            job = orchestrator.create_job({
                "workflow": "event_speakers",
                "source_url": source_url,
                "enrich_missing": True,
                "search_provider": "ddgs",
            })
            with patch("speedy_scraper.event_speakers.ProviderRegistry", return_value=registry):
                completed = orchestrator.run_job(job.id)

            self.assertEqual(completed.status, JobStatus.COMPLETED)
            artifacts = repository.list_artifacts(job.id, "event_speakers")
            self.assertEqual(len(artifacts), 1)
            rows = artifacts[0]["payload"]
            self.assertEqual(len(rows), 2)
            self.assertEqual([row["Match Status"] for row in rows], ["provided", "matched"])
            self.assertEqual(repository.get_job(job.id).checkpoint["event_speakers"]["phase"], "completed")
            self.assertEqual(repository.metrics(job.id)["speakers_extracted"], 2)

            csv_bytes, media_type, csv_name = export_artifacts(repository, job.id, "csv")
            self.assertEqual(media_type, "text/csv")
            self.assertTrue(csv_name.endswith("-event-speakers.csv"))
            self.assertEqual(csv_bytes.decode().splitlines()[0].split(","), EVENT_SPEAKER_EXPORT_COLUMNS)
            json_bytes, _, _ = export_artifacts(repository, job.id, "json")
            self.assertEqual(len(json.loads(json_bytes)), 2)
            xlsx_bytes, _, _ = export_artifacts(repository, job.id, "xlsx")
            workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
            self.assertEqual(workbook.sheetnames, ["Speakers"])

    def test_pause_checkpoint_and_resume_without_refetching(self):
        with tempfile.TemporaryDirectory() as directory:
            config, repository, orchestrator = self._services(directory)
            source_url = "https://events.example/speakers"
            fetcher = Mock()
            fetcher.fetch.return_value = _flight_html([
                _row("Jane Doe", linkedin=None),
                _row("Bob Singh", company="Beta", linkedin=None),
            ])
            engine = EventSpeakerEngine(config, repository, fetcher=fetcher)
            job = repository.create_job("event_speakers", {
                "workflow": "event_speakers",
                "source_url": source_url,
                "enrich_missing": True,
                "search_provider": "ddgs",
            })
            provider = Mock()

            def pause_after_first(_request):
                repository.transition(job.id, JobStatus.PAUSE_REQUESTED, outcome="Pause requested")
                return SearchPage((), "ddgs", 1)

            provider.search.side_effect = pause_after_first
            registry = Mock()
            registry.get.return_value = provider
            with patch("speedy_scraper.event_speakers.ProviderRegistry", return_value=registry):
                paused = engine.run(job.id)
            self.assertEqual(paused.status, JobStatus.PAUSED)
            self.assertEqual(repository.get_job(job.id).checkpoint["event_speakers"]["speaker_index"], 1)
            self.assertEqual(fetcher.fetch.call_count, 1)

            repository.transition(job.id, JobStatus.QUEUED, outcome="Queued")
            provider.search.side_effect = None
            provider.search.return_value = SearchPage((), "ddgs", 1)
            with patch("speedy_scraper.event_speakers.ProviderRegistry", return_value=registry):
                completed = engine.run(job.id)
            self.assertEqual(completed.status, JobStatus.COMPLETED)
            self.assertEqual(fetcher.fetch.call_count, 1)

    def test_provider_no_results_exception_becomes_not_found(self):
        with tempfile.TemporaryDirectory() as directory:
            config, repository, _orchestrator = self._services(directory)
            fetcher = Mock()
            fetcher.fetch.return_value = _flight_html([_row("Jane Doe", linkedin=None)])
            engine = EventSpeakerEngine(config, repository, fetcher=fetcher)
            job = repository.create_job("event_speakers", {
                "workflow": "event_speakers",
                "source_url": "https://events.example/speakers",
                "enrich_missing": True,
                "search_provider": "ddgs",
            })
            provider = Mock()
            provider.search.side_effect = RuntimeError("No results found.")
            registry = Mock()
            registry.get.return_value = provider
            with patch("speedy_scraper.event_speakers.ProviderRegistry", return_value=registry):
                completed = engine.run(job.id)

            self.assertEqual(completed.status, JobStatus.COMPLETED)
            row = repository.list_artifacts(job.id, "event_speakers")[0]["payload"][0]
            self.assertEqual(row["Match Status"], "not_found")
            self.assertEqual(repository.metrics(job.id)["speaker_search_empty"], 2)

    def test_cancel_request_stops_before_fetch(self):
        with tempfile.TemporaryDirectory() as directory:
            config, repository, _orchestrator = self._services(directory)
            fetcher = Mock()
            engine = EventSpeakerEngine(config, repository, fetcher=fetcher)
            job = repository.create_job("event_speakers", {
                "workflow": "event_speakers",
                "source_url": "https://events.example/speakers",
                "enrich_missing": False,
                "search_provider": "ddgs",
            })
            repository.transition(job.id, JobStatus.RUNNING)
            repository.transition(job.id, JobStatus.CANCEL_REQUESTED)

            cancelled = engine.run(job.id)

            self.assertEqual(cancelled.status, JobStatus.CANCELLED)
            fetcher.fetch.assert_not_called()

    def test_request_normalization_rejects_invalid_provider_and_source(self):
        with tempfile.TemporaryDirectory() as directory:
            _config, _repository, orchestrator = self._services(directory)
            with self.assertRaisesRegex(ValueError, "ddgs or brave"):
                orchestrator.create_job({
                    "workflow": "event_speakers",
                    "source_url": "https://events.example/speakers",
                    "search_provider": "google",
                })
            with self.assertRaises(UnsafeSourceUrlError):
                orchestrator.create_job({
                    "workflow": "event_speakers",
                    "source_url": "file:///tmp/speakers",
                })


if __name__ == "__main__":
    unittest.main()
