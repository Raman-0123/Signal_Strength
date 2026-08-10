"""Stable CSV, JSON, and Excel exports."""

from __future__ import annotations

import io
import json
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from speedy_scraper.domain import ResultKind

LEAD_EXPORT_COLUMNS = ["Name", "Designation", "Company", "Location"]
EVENT_SPEAKER_EXPORT_COLUMNS = [
    "Name",
    "Designation",
    "Company",
    "Country",
    "LinkedIn ID",
    "LinkedIn URL",
    "Match Status",
    "Confidence",
    "Match Evidence",
    "Source URL",
]


def _excel_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            sheet_name = str(name)[:31]
            frame.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            fill = PatternFill("solid", fgColor="1F4E78")
            for cell in worksheet[1]:
                cell.fill = fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for index, column in enumerate(frame.columns, start=1):
                values = [str(column), *[str(value) for value in frame[column].dropna().head(500)]]
                worksheet.column_dimensions[get_column_letter(index)].width = min(
                    max(max((len(value) for value in values), default=10) + 2, 12), 48,
                )
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
    return buffer.getvalue()


def export_leads(repository, job_id: str, fmt: str, kind: ResultKind) -> tuple[bytes, str, str]:
    records = repository.list_results(job_id, kind)
    rows = [record.as_export_row() for record in records]
    frame = pd.DataFrame(rows, columns=LEAD_EXPORT_COLUMNS)
    normalized = fmt.lower()
    suffix = "strict" if kind == ResultKind.STRICT else "qualified"
    if normalized == "csv":
        return frame.to_csv(index=False).encode("utf-8"), "text/csv", f"{job_id}-{suffix}.csv"
    if normalized == "json":
        return json.dumps(rows, ensure_ascii=False, indent=2).encode("utf-8"), "application/json", f"{job_id}-{suffix}.json"
    if normalized == "xlsx":
        qualified = pd.DataFrame(
            [record.as_export_row() for record in repository.list_results(job_id, ResultKind.QUALIFIED)],
            columns=LEAD_EXPORT_COLUMNS,
        )
        strict = pd.DataFrame(
            [record.as_export_row() for record in repository.list_results(job_id, ResultKind.STRICT)],
            columns=LEAD_EXPORT_COLUMNS,
        )
        return _excel_bytes({"All Qualified POCs": qualified, "Strict Matches": strict}), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"{job_id}-leads.xlsx"
    raise ValueError(f"Unsupported export format: {fmt}")


def export_artifacts(repository, job_id: str, fmt: str) -> tuple[bytes, str, str]:
    artifacts = repository.list_artifacts(job_id)
    by_type = {artifact["artifact_type"]: artifact["payload"] for artifact in artifacts}
    if "event_speakers" in by_type:
        event_rows = list(by_type["event_speakers"])
        frame = pd.DataFrame(event_rows, columns=EVENT_SPEAKER_EXPORT_COLUMNS)
        if fmt == "csv":
            return (
                frame.to_csv(index=False).encode("utf-8"),
                "text/csv",
                f"{job_id}-event-speakers.csv",
            )
        if fmt == "json":
            return (
                json.dumps(event_rows, ensure_ascii=False, indent=2).encode("utf-8"),
                "application/json",
                f"{job_id}-event-speakers.json",
            )
        if fmt == "xlsx":
            return (
                _excel_bytes({"Speakers": frame}),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"{job_id}-event-speakers.xlsx",
            )
        raise ValueError(f"Unsupported export format: {fmt}")
    if "competitor_leads" in by_type:
        summaries = by_type.get("company_summaries", {})
        csev_rows = []
        for lead in by_type["competitor_leads"]:
            summary = summaries.get(lead.get("Company", ""), {})
            csev_rows.append({
                "Name": lead.get("Full_Name", ""),
                "COMPANY": lead.get("Company", ""),
                "DESIGNATION": lead.get("Designation", ""),
                "PHONE NUMBER": "",
                "LinkedIn Id": lead.get("LinkedIn_URL", ""),
                "Competitor Source": lead.get("Competitor_Source", ""),
                "Event Signal": lead.get("Event_Signal", ""),
                "Event Probability": lead.get("Event_Probability", ""),
                "Signal Confidence": lead.get("Signal_Confidence", ""),
                "Signal Evidence": lead.get("Signal_Evidence", ""),
                "Location Verified": lead.get("Location_Verified", ""),
                "What Company Does": summary.get("description", "—"),
                "Net Profit": summary.get("net_profit", "—"),
                "Matched Parameters": lead.get("Matched_Parameters", ""),
            })
        csev = pd.DataFrame(csev_rows)
        if fmt == "csv":
            return csev.to_csv(index=False).encode(), "text/csv", f"{job_id}-CSEV.csv"
        if fmt == "xlsx":
            return _excel_bytes({"CSEV": csev}), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"{job_id}-CSEV.xlsx"
    rows: list[dict[str, Any]] = []
    for artifact in artifacts:
        payload = artifact["payload"]
        if isinstance(payload, list):
            rows.extend(item if isinstance(item, dict) else {"value": item} for item in payload)
        else:
            rows.append(payload if isinstance(payload, dict) else {"value": payload})
    if fmt == "json":
        return json.dumps(artifacts, ensure_ascii=False, indent=2, default=str).encode(), "application/json", f"{job_id}-artifacts.json"
    frame = pd.DataFrame(rows)
    if fmt == "csv":
        return frame.to_csv(index=False).encode(), "text/csv", f"{job_id}-artifacts.csv"
    if fmt == "xlsx":
        return _excel_bytes({"Results": frame}), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"{job_id}-artifacts.xlsx"
    raise ValueError(f"Unsupported export format: {fmt}")
