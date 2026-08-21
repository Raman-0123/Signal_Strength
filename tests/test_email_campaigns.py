from __future__ import annotations

from datetime import UTC, datetime, timedelta
from pathlib import Path
from urllib.parse import parse_qs, urlparse
from zoneinfo import ZoneInfo

import pandas as pd
import pytest
from cryptography.fernet import Fernet
from openpyxl import load_workbook
from sqlalchemy import select

from speedy_scraper.campaign_db import init_database, reset_engine_cache, session_scope
from speedy_scraper.campaign_models import (
    Campaign,
    CampaignLink,
    CampaignRecipient,
    CampaignStatus,
    EmailAccount,
    OAuthState,
    RecipientStatus,
    ResponseClassification,
    SuppressionEntry,
)
from speedy_scraper.campaign_worker import (
    _process_mailbox_message,
    process_one_due,
    recover_stale_leases,
)
from speedy_scraper.email_campaigns import (
    CampaignSchedule,
    RecipientCandidate,
    campaign_history_frame,
    campaign_workbook,
    candidates_from_frame,
    configure_campaign,
    create_campaign,
    pause_campaign,
    recipient_context,
    render_template,
    schedule_campaign,
    strategic_send_slots,
    validate_templates,
)
from speedy_scraper.email_sender import (
    SenderError,
    SendResult,
    build_email_message,
    build_gmail_authorization_url,
    decrypt_credentials,
    encrypt_credentials,
)


@pytest.fixture
def campaign_db(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> str:
    reset_engine_cache()
    path = tmp_path / "campaigns.db"
    monkeypatch.delenv("DATABASE_URL", raising=False)
    monkeypatch.setenv("CAMPAIGN_SQLITE_PATH", str(path))
    monkeypatch.setenv("PUBLIC_API_URL", "http://127.0.0.1:8000")
    init_database()
    yield str(path)
    reset_engine_cache()


def _draft(session, count: int = 2):
    candidates = [
        RecipientCandidate(
            email=f"person{index}@example.com",
            full_name=f"Person {index}",
            first_name=f"Person{index}",
            company="Signal Labs",
            designation="Leader",
            source_record_id=str(index),
        )
        for index in range(count)
    ]
    campaign, rejected = create_campaign(session, name="Strategic invite", candidates=candidates)
    assert rejected == []
    configure_campaign(
        campaign,
        subject_template="A note for {{first_name}}",
        html_template='<p>Hi {{first_name}} at <a href="https://example.com">{{company}}</a></p>',
        text_template="Hi {{first_name}} at {{company}}",
        sender_name="Raman",
        reply_to="raman@example.com",
        organization_name="Signal",
        organization_address="Mumbai, India",
    )
    return campaign


def _schedule(session, campaign, *, daily_target: int = 100):
    account = EmailAccount(
        email="sender@gmail.com",
        encrypted_credentials="not-used-by-tests",
    )
    session.add(account)
    session.flush()
    schedule_campaign(
        session,
        campaign,
        account=account,
        start_at=datetime.now(UTC) - timedelta(minutes=1),
        schedule=CampaignSchedule(
            timezone="Asia/Kolkata",
            allowed_weekdays=(0, 1, 2, 3, 4, 5, 6),
            window_start_hour=0,
            window_end_hour=24,
            interval_min_minutes=1,
            interval_max_minutes=1,
            daily_target=daily_target,
        ),
        public_api_url="http://127.0.0.1:8000",
    )
    session.flush()
    return account


def test_recipient_intake_requires_safe_status_and_rejects_invalid_duplicates():
    frame = pd.DataFrame(
        [
            {"POC": "Asha Rao", "Work Email": "ASHA@example.com", "Decision": "SAFE TO CONTACT"},
            {"POC": "Duplicate", "Work Email": "asha@example.com", "Decision": "SAFE TO CONTACT"},
            {"POC": "Unsafe", "Work Email": "unsafe@example.com", "Decision": "ALREADY CONTACTED"},
            {"POC": "Broken", "Work Email": "not-an-email", "Decision": "SAFE TO CONTACT"},
        ]
    )

    accepted, rejected = candidates_from_frame(
        frame,
        {"full_name": "POC", "email": "Work Email", "decision": "Decision"},
    )

    assert [item.email for item in accepted] == ["asha@example.com"]
    assert {item.reason for item in rejected} == {
        "duplicate_email_in_source",
        "not_safe_to_contact",
        "invalid_or_blank_email",
    }


def test_direct_upload_requires_explicit_safe_confirmation():
    frame = pd.DataFrame([{"Name": "Asha", "Email": "asha@example.com"}])

    accepted, rejected = candidates_from_frame(frame, {"full_name": "Name", "email": "Email"})
    confirmed, confirmed_rejected = candidates_from_frame(
        frame, {"full_name": "Name", "email": "Email"}, confirmed_safe=True
    )

    assert accepted == []
    assert rejected[0].reason == "safe_status_not_confirmed"
    assert len(confirmed) == 1
    assert confirmed_rejected == []


def test_personalization_is_deterministic_and_missing_values_block_scheduling():
    candidate = RecipientCandidate(
        email="asha@example.com",
        full_name="Asha Rao",
        company="R&D <Labs>",
        source_data={"Event Theme": "Customer trust"},
    )
    context = recipient_context(candidate)

    assert context["first_name"] == "Asha"
    assert context["event_theme"] == "Customer trust"
    assert render_template("Hi {{first_name}}", context, html_mode=False) == "Hi Asha"
    assert "R&amp;D" in render_template("<p>{{company}}</p>", context, html_mode=True)
    assert validate_templates([candidate], "{{first_name}}", "{{designation}}", "") == {
        "asha@example.com": ["designation"]
    }
    with pytest.raises(ValueError, match="simple placeholders"):
        render_template("{{ name.upper() }}", context, html_mode=False)


def test_strategic_schedule_has_no_campaign_size_cap_and_respects_daily_target():
    start = datetime(2026, 8, 18, 10, 0, tzinfo=ZoneInfo("Asia/Kolkata"))
    schedule = CampaignSchedule(
        daily_target=100,
        interval_min_minutes=1,
        interval_max_minutes=1,
    )

    slots = strategic_send_slots(start, 650, schedule, seed="large-campaign")

    assert len(slots) == 650
    local_dates = [slot.astimezone(ZoneInfo("Asia/Kolkata")).date() for slot in slots]
    assert max(local_dates.count(day) for day in set(local_dates)) <= 100
    assert all(slot.astimezone(ZoneInfo("Asia/Kolkata")).weekday() < 5 for slot in slots)


def test_create_schedule_snapshots_content_and_tracking_links(campaign_db: str):
    with session_scope() as session:
        campaign = _draft(session, count=1)
        _schedule(session, campaign)
        recipient = session.scalar(
            select(CampaignRecipient).where(CampaignRecipient.campaign_id == campaign.id)
        )
        links = list(
            session.scalars(select(CampaignLink).where(CampaignLink.recipient_id == recipient.id))
        )

        assert campaign.status == CampaignStatus.SCHEDULED.value
        assert recipient.status == RecipientStatus.QUEUED.value
        assert recipient.rendered_subject == "A note for Person0"
        assert "/track/open/" in recipient.rendered_html
        assert "/unsubscribe/" in recipient.rendered_html
        assert len(links) == 1
        assert links[0].destination_url == "https://example.com"


def test_rescheduling_a_paused_campaign_never_requeues_already_sent_people(campaign_db: str):
    with session_scope() as session:
        campaign = _draft(session, count=2)
        account = _schedule(session, campaign)
        recipients = list(
            session.scalars(
                select(CampaignRecipient)
                .where(CampaignRecipient.campaign_id == campaign.id)
                .order_by(CampaignRecipient.created_at)
            )
        )
        recipients[0].status = RecipientStatus.SENT.value
        pause_campaign(campaign)
        schedule_campaign(
            session,
            campaign,
            account=account,
            start_at=datetime.now(UTC),
            schedule=CampaignSchedule(
                timezone="Asia/Kolkata",
                allowed_weekdays=(0, 1, 2, 3, 4, 5, 6),
                window_start_hour=0,
                window_end_hour=24,
                interval_min_minutes=1,
                interval_max_minutes=1,
                daily_target=100,
            ),
            public_api_url="http://127.0.0.1:8000",
        )

        assert recipients[0].status == RecipientStatus.SENT.value
        assert recipients[1].status == RecipientStatus.QUEUED.value


def test_suppression_is_rechecked_when_draft_is_created(campaign_db: str):
    with session_scope() as session:
        session.add(
            SuppressionEntry(
                normalized_email="blocked@example.com", reason="unsubscribe", source="test"
            )
        )
        campaign, rejected = create_campaign(
            session,
            name="Suppression",
            candidates=[RecipientCandidate(email="blocked@example.com", full_name="Blocked")],
        )

        assert rejected[0].reason == "suppressed"
        assert campaign.recipients == []


def test_campaign_history_blocks_duplicates_but_recontact_override_is_respected(
    campaign_db: str,
):
    candidate = RecipientCandidate(email="asha@example.com", full_name="Asha")
    with session_scope() as session:
        first, _ = create_campaign(session, name="First", candidates=[candidate])
        original = session.scalar(
            select(CampaignRecipient).where(CampaignRecipient.campaign_id == first.id)
        )
        original.status = RecipientStatus.SENT.value
        second, rejected = create_campaign(session, name="Second", candidates=[candidate])

        assert rejected[0].reason == "already_in_campaign_history"
        assert second.recipients == []

        original.response_classification = ResponseClassification.RECONTACT_ALLOWED.value
        third, third_rejected = create_campaign(session, name="Third", candidates=[candidate])

        assert third_rejected == []
        assert len(third.recipients) == 1


class _FakeSender:
    sent = 0

    def __init__(self, _account):
        pass

    def find_sent_message(self, _message_id):
        return None

    def send(self, _message):
        type(self).sent += 1
        return SendResult(message_id="gmail-1", thread_id="thread-1")


class _QuotaSender(_FakeSender):
    def send(self, _message):
        raise SenderError("dailyLimitExceeded", retryable=True, quota_limited=True)


class _RevokedSender(_FakeSender):
    def send(self, _message):
        raise SenderError(
            "invalid_grant",
            retryable=False,
            authentication_failed=True,
        )


def test_worker_sends_once_and_reconciles_idempotently(campaign_db: str):
    _FakeSender.sent = 0
    with session_scope() as session:
        campaign = _draft(session, count=1)
        _schedule(session, campaign)
        recipient = session.scalar(
            select(CampaignRecipient).where(CampaignRecipient.campaign_id == campaign.id)
        )
        recipient.scheduled_at = datetime.now(UTC) - timedelta(minutes=1)
        recipient_id = recipient.id

    assert process_one_due(worker_id="test-worker", sender_factory=_FakeSender) is True
    assert process_one_due(worker_id="test-worker", sender_factory=_FakeSender) is False
    with session_scope() as session:
        recipient = session.get(CampaignRecipient, recipient_id)
        assert recipient.status == RecipientStatus.SENT.value
        assert recipient.gmail_message_id == "gmail-1"
        assert _FakeSender.sent == 1


def test_quota_failure_stays_queued_for_a_later_window(campaign_db: str):
    with session_scope() as session:
        campaign = _draft(session, count=1)
        _schedule(session, campaign)
        recipient = session.scalar(
            select(CampaignRecipient).where(CampaignRecipient.campaign_id == campaign.id)
        )
        recipient.scheduled_at = datetime.now(UTC) - timedelta(minutes=1)
        recipient_id = recipient.id

    before = datetime.now(UTC)
    assert process_one_due(worker_id="quota-worker", sender_factory=_QuotaSender) is True
    with session_scope() as session:
        recipient = session.get(CampaignRecipient, recipient_id)
        assert recipient.status == RecipientStatus.QUEUED.value
        assert recipient.scheduled_at.replace(tzinfo=recipient.scheduled_at.tzinfo or UTC) > before
        assert "dailyLimitExceeded" in recipient.last_error


def test_revoked_gmail_authorization_pauses_campaign_without_losing_queue(campaign_db: str):
    with session_scope() as session:
        campaign = _draft(session, count=1)
        account = _schedule(session, campaign)
        recipient = session.scalar(select(CampaignRecipient))
        recipient.scheduled_at = datetime.now(UTC) - timedelta(minutes=1)
        campaign_id, account_id, recipient_id = campaign.id, account.id, recipient.id

    assert process_one_due(worker_id="auth-worker", sender_factory=_RevokedSender) is True
    with session_scope() as session:
        campaign = session.get(Campaign, campaign_id)
        account = session.get(EmailAccount, account_id)
        recipient = session.get(CampaignRecipient, recipient_id)
        assert campaign.status == CampaignStatus.PAUSED.value
        assert account.active is False
        assert recipient.status == RecipientStatus.QUEUED.value


def test_expired_worker_lease_is_recovered_without_discarding_recipient(campaign_db: str):
    with session_scope() as session:
        campaign = _draft(session, count=1)
        _schedule(session, campaign)
        recipient = session.scalar(
            select(CampaignRecipient).where(CampaignRecipient.campaign_id == campaign.id)
        )
        recipient.status = RecipientStatus.SENDING.value
        recipient.leased_until = datetime.now(UTC) - timedelta(minutes=1)
        recipient.lease_owner = "dead-worker"

    with session_scope() as session:
        assert recover_stale_leases(session) == 1
        recipient = session.scalar(select(CampaignRecipient))
        assert recipient.status == RecipientStatus.QUEUED.value
        assert "Sent is checked before retry" in recipient.last_error


def test_mailbox_history_classifies_replies_and_bounces(campaign_db: str):
    class _Request:
        def __init__(self, value):
            self.value = value

        def execute(self):
            return self.value

    class _Messages:
        def __init__(self, values):
            self.values = values

        def get(self, **kwargs):
            return _Request(self.values[kwargs["id"]])

    class _Users:
        def __init__(self, values):
            self.values = values

        def messages(self):
            return _Messages(self.values)

    class _Service:
        def __init__(self, values):
            self.values = values

        def users(self):
            return _Users(self.values)

    with session_scope() as session:
        campaign = _draft(session, count=2)
        account = _schedule(session, campaign)
        recipients = list(
            session.scalars(
                select(CampaignRecipient).order_by(CampaignRecipient.created_at)
            )
        )
        recipients[0].status = RecipientStatus.SENT.value
        recipients[0].gmail_thread_id = "reply-thread"
        recipients[1].status = RecipientStatus.SENT.value
        recipients[1].gmail_thread_id = "bounce-thread"
        values = {
            "reply-message": {
                "threadId": "reply-thread",
                "payload": {
                    "headers": [
                        {"name": "From", "value": "Person <person0@example.com>"},
                        {"name": "Subject", "value": "Re: invitation"},
                    ]
                },
            },
            "bounce-message": {
                "threadId": "bounce-thread",
                "payload": {
                    "headers": [
                        {"name": "From", "value": "Mailer-Daemon@example.com"},
                        {"name": "Subject", "value": "Delivery Status Notification"},
                    ]
                },
            },
        }
        fake_sender = type("FakeGmail", (), {"service": _Service(values)})()

        assert _process_mailbox_message(
            session, account, fake_sender, "reply-message"
        ) == 1
        assert _process_mailbox_message(
            session, account, fake_sender, "bounce-message"
        ) == 1

        assert recipients[0].status == RecipientStatus.REPLIED.value
        assert recipients[1].status == RecipientStatus.BOUNCED.value
        assert session.scalar(
            select(SuppressionEntry).where(
                SuppressionEntry.normalized_email == recipients[1].normalized_email
            )
        ) is not None


def test_history_and_audit_exports_preserve_campaign_evidence(campaign_db: str, tmp_path: Path):
    with session_scope() as session:
        campaign = _draft(session, count=1)
        _schedule(session, campaign)
        recipient = session.scalar(select(CampaignRecipient))
        recipient.status = RecipientStatus.SENT.value
        recipient.sent_at = datetime.now(UTC)

        history = campaign_history_frame(session)
        workbook = campaign_workbook(session, campaign)

    assert history.iloc[0]["Status"] == "Contacted"
    target = tmp_path / "audit.xlsx"
    target.write_bytes(workbook)
    assert load_workbook(target, read_only=True).sheetnames == [
        "Recipients",
        "Message Events",
        "Replies",
        "Failures",
        "Unsubscribes",
        "Configuration",
    ]


def test_credentials_are_encrypted_and_messages_include_suppression_headers(monkeypatch):
    monkeypatch.setenv("TOKEN_ENCRYPTION_KEY", Fernet.generate_key().decode())
    encrypted = encrypt_credentials('{"refresh_token":"secret"}')

    message = build_email_message(
        sender_email="sender@gmail.com",
        sender_name="Raman",
        recipient_email="asha@example.com",
        subject="Invitation",
        text_body="Hello",
        html_body="<p>Hello</p>",
        reply_to="reply@example.com",
        rfc_message_id="<stable-id@signal.local>",
        unsubscribe_url="https://api.example.com/unsubscribe/token",
    )

    assert "secret" not in encrypted
    assert decrypt_credentials(encrypted) == '{"refresh_token":"secret"}'
    assert message["Message-ID"] == "<stable-id@signal.local>"
    assert message["List-Unsubscribe-Post"] == "List-Unsubscribe=One-Click"
    assert "unsubscribe/token" in str(message["List-Unsubscribe"])


def test_oauth_authorization_uses_a_short_lived_persisted_state(
    campaign_db: str, monkeypatch: pytest.MonkeyPatch
):
    monkeypatch.setenv("GOOGLE_CLIENT_ID", "client-id.apps.googleusercontent.com")
    monkeypatch.setenv("GOOGLE_CLIENT_SECRET", "client-secret")
    monkeypatch.setenv(
        "GOOGLE_OAUTH_REDIRECT_URI",
        "http://127.0.0.1:8000/oauth/gmail/callback",
    )
    with session_scope() as session:
        url = build_gmail_authorization_url(session)
        state = parse_qs(urlparse(url).query)["state"][0]
        stored = session.get(OAuthState, state)

        assert stored is not None
        assert stored.used_at is None
        assert stored.expires_at > datetime.now(UTC).replace(tzinfo=None)
