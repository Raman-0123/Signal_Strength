import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from speedy_scraper.background_jobs import clear_stop, create_job, job_is_stale, read_status
from speedy_scraper.company_pocs import build_company_poc_tasks, run_company_poc_job
from speedy_scraper.lead_job import load_lead_job_checkpoint, run_lead_job
from speedy_scraper.models import SearchPage, SearchResult
from speedy_scraper.sources import SourceError
from speedy_scraper.url_people_job import load_checkpoint_speakers, run_url_people_job


class StopAfterFirstUrlSearch:
    name = "fake"

    def __init__(self, job_dir: Path):
        self.job_dir = job_dir

    def search(self, query, *, max_results, headless=True):
        name = "Asha Rao" if "Asha Rao" in query else "Bina Shah"
        slug = "asha-rao" if name == "Asha Rao" else "bina-shah"
        company = "Razorpay" if name == "Asha Rao" else "PhonePe"
        (self.job_dir / "stop.requested").touch()
        return [
            SearchResult(
                title=f"{name} - CTO - {company} | LinkedIn",
                body=f"{name} is Chief Technology Officer at {company}.",
                href=f"https://www.linkedin.com/in/{slug}/",
                source="fake",
                query=query,
            )
        ]


class ResumeUrlSearch:
    name = "fake"

    def __init__(self):
        self.queries = []

    def search(self, query, *, max_results, headless=True):
        self.queries.append(query)
        return [
            SearchResult(
                title="Bina Shah - CTO - PhonePe | LinkedIn",
                body="Bina Shah is Chief Technology Officer at PhonePe.",
                href="https://www.linkedin.com/in/bina-shah/",
                source="fake",
                query=query,
            )
        ]


def test_url_job_stops_and_resumes_after_checkpoint(tmp_path: Path):
    job_dir = create_job(
        "url_people",
        {"source_url": "https://example.com/people", "sources": ["fake"]},
        jobs_root=tmp_path,
    )
    html = json.dumps(
        {
            "people": [
                {"id": "1", "name": "Asha Rao", "designation": "CTO", "company": "Razorpay"},
                {"id": "2", "name": "Bina Shah", "designation": "CTO", "company": "PhonePe"},
            ]
        }
    )
    first_source = StopAfterFirstUrlSearch(job_dir)
    run_url_people_job(
        job_dir,
        fetcher=lambda _url: html,
        source_builder=lambda _names: [first_source],
    )
    speakers, next_index = load_checkpoint_speakers(job_dir)
    assert read_status(job_dir)["state"] == "paused"
    assert next_index == 1
    assert speakers[0].linkedin_url.endswith("/asha-rao/")

    clear_stop(job_dir)
    resume_source = ResumeUrlSearch()
    run_url_people_job(job_dir, source_builder=lambda _names: [resume_source])
    speakers, next_index = load_checkpoint_speakers(job_dir)
    assert read_status(job_dir)["state"] == "completed"
    assert next_index == 2
    assert len(resume_source.queries) == 1
    assert "Bina Shah" in resume_source.queries[0]
    assert speakers[1].linkedin_url.endswith("/bina-shah/")


class StopAfterFirstPocSearch:
    name = "fake"

    def __init__(self, job_dir: Path, stop: bool):
        self.job_dir = job_dir
        self.stop = stop
        self.queries = []

    def search(self, query, *, max_results, headless=True):
        self.queries.append(query)
        if "Razorpay" in query:
            name, company, slug = "Asha Rao", "Razorpay", "asha-rao"
        else:
            name, company, slug = "Bina Shah", "PhonePe", "bina-shah"
        if self.stop:
            (self.job_dir / "stop.requested").touch()
        return [
            SearchResult(
                title=f"{name} - Chief Technology Officer - {company} | LinkedIn",
                body=f"{company} CTO profile",
                href=f"https://www.linkedin.com/in/{slug}/",
                source="fake",
                query=query,
            )
        ]


def test_company_poc_job_uses_company_role_pairs_and_resumes(tmp_path: Path):
    tasks = build_company_poc_tasks(["Razorpay", "PhonePe"], ["CTO"])
    assert len(tasks) == 2
    assert all("site:linkedin.com/in" in task["query"] for task in tasks)

    job_dir = create_job(
        "company_pocs",
        {
            "companies": ["Razorpay", "PhonePe"],
            "designations": ["CTO"],
            "sources": ["fake"],
            "target_count": 10,
        },
        jobs_root=tmp_path,
    )
    first = StopAfterFirstPocSearch(job_dir, stop=True)
    pocs = run_company_poc_job(job_dir, source_builder=lambda _names: [first])
    assert read_status(job_dir)["state"] == "paused"
    assert [poc.company for poc in pocs] == ["Razorpay"]

    clear_stop(job_dir)
    second = StopAfterFirstPocSearch(job_dir, stop=False)
    pocs = run_company_poc_job(job_dir, source_builder=lambda _names: [second])
    final_status = read_status(job_dir)
    assert final_status["state"] == "completed"
    assert final_status["searches_completed"] == 2
    assert final_status["searches_total"] == 2
    assert len(second.queries) == 1
    assert "PhonePe" in second.queries[0]
    assert {poc.company for poc in pocs} == {"Razorpay", "PhonePe"}


def test_company_poc_query_passes_expand_search_plan_for_user_selected_depth():
    tasks = build_company_poc_tasks(
        ["Razorpay"],
        ["CTO"],
        ["Singapore"],
        query_passes=3,
    )
    assert len(tasks) == 3
    assert [task["query_variant"] for task in tasks] == ["1", "2", "3"]
    assert len({task["query"] for task in tasks}) == 3


class WrongRoleSource:
    name = "fake"

    def search(self, query, *, max_results, headless=True):
        return [
            SearchResult(
                title="Balaji Bandlapalli - Principal Engineer II - Razorpay | LinkedIn",
                body="Engineering leader at Razorpay.",
                href="https://www.linkedin.com/in/balajibandlapalli/",
                source="fake",
                query=query,
            )
        ]


def test_company_poc_job_rejects_company_match_with_wrong_designation(tmp_path: Path):
    job_dir = create_job(
        "company_pocs",
        {
            "companies": ["Razorpay"],
            "designations": ["Head of Customer Success"],
            "sources": ["fake"],
            "target_count": 10,
        },
        jobs_root=tmp_path,
    )
    pocs = run_company_poc_job(job_dir, source_builder=lambda _names: [WrongRoleSource()])
    assert pocs == []


class BlendedEvidenceSource:
    name = "fake"

    def search(self, query, *, max_results, headless=True):
        return [
            SearchResult(
                title="Balaji Bandlapalli - Principal Engineer II - Razorpay | LinkedIn",
                body=(
                    "Engineering leader at Razorpay. Related result: Jane Doe is the "
                    "Chief Information Officer at another company."
                ),
                href="https://www.linkedin.com/in/balajibandlapalli/",
                source="fake",
                query=query,
            ),
            SearchResult(
                title="Khanjan Desaai - Payments Product - Booking.com | LinkedIn",
                body="Previously built a product at Razorpay. Related profiles include CX leaders.",
                href="https://www.linkedin.com/in/khanjandesai8/",
                source="fake",
                query=query,
            ),
        ]


def test_company_poc_job_rejects_role_or_company_found_only_in_blended_evidence(tmp_path: Path):
    job_dir = create_job(
        "company_pocs",
        {
            "companies": ["Razorpay"],
            "designations": ["CIO"],
            "sources": ["fake"],
            "target_count": 10,
        },
        jobs_root=tmp_path,
    )

    pocs = run_company_poc_job(job_dir, source_builder=lambda _names: [BlendedEvidenceSource()])

    assert pocs == []


class ChallengeDdgsSource:
    name = "ddgs"

    def search(self, query, *, max_results, headless=True):
        raise SourceError("DDG challenge", disable_source=True, challenge=True)


class SuccessfulGoogleSource:
    name = "google_browser"

    def search(self, query, *, max_results, headless=True):
        return [
            SearchResult(
                title="Meera Iyer - Senior Director of Customer Experience - MUFG | LinkedIn",
                body="Meera Iyer is a Senior Director of Customer Experience at MUFG in Singapore.",
                href="https://www.linkedin.com/in/meera-iyer/",
                source="google_browser",
                query=query,
            )
        ]


def test_company_poc_retries_ddg_failure_with_google_and_keeps_warning_telemetry(tmp_path: Path):
    job_dir = create_job(
        "company_pocs",
        {
            "companies": ["MUFG Bank Singapore"],
            "designations": ["Senior Director"],
            "locations": ["Singapore"],
            "sources": ["ddgs"],
            "retry_attempts": 1,
            "target_count": 10,
        },
        jobs_root=tmp_path,
    )
    pocs = run_company_poc_job(
        job_dir,
        source_builder=lambda _names: [ChallengeDdgsSource(), SuccessfulGoogleSource()],
    )
    status = read_status(job_dir)
    assert len(pocs) == 1
    assert status["state"] == "completed_with_warnings"
    assert status["provider_outcomes"]["ddgs"]["challenges"] >= 1
    assert status["provider_outcomes"]["google_browser"]["results"] >= 1
    assert status["failed_searches"] == 0


class AlwaysFailSource:
    def __init__(self, name):
        self.name = name

    def search(self, query, *, max_results, headless=True):
        raise SourceError(f"{self.name} unavailable", disable_source=True)


def test_company_poc_all_provider_failures_finish_with_warnings_and_retry_queue(tmp_path: Path):
    job_dir = create_job(
        "company_pocs",
        {
            "companies": ["MUFG"],
            "designations": ["Senior Director"],
            "sources": ["google_browser", "ddgs"],
            "retry_attempts": 1,
        },
        jobs_root=tmp_path,
    )
    run_company_poc_job(
        job_dir,
        source_builder=lambda _names: [AlwaysFailSource("google_browser"), AlwaysFailSource("ddgs")],
    )
    status = read_status(job_dir)
    checkpoint = json.loads((job_dir / "checkpoint.json").read_text())
    assert status["state"] == "completed_with_warnings"
    assert status["matched"] == 0
    assert status["failed_searches"] > 0
    assert checkpoint["warning_state"] == "completed_with_warnings"


def test_company_poc_warning_checkpoint_can_retry_same_provider_after_recovery(tmp_path: Path):
    job_dir = create_job(
        "company_pocs",
        {
            "companies": ["MUFG Bank Singapore"],
            "designations": ["Senior Director"],
            "sources": ["google_browser"],
            "retry_attempts": 1,
        },
        jobs_root=tmp_path,
    )
    run_company_poc_job(
        job_dir,
        source_builder=lambda _names: [AlwaysFailSource("google_browser")],
    )
    config = json.loads((job_dir / "config.json").read_text())
    config["retry_failed_searches"] = True
    config["browser_headless"] = False
    (job_dir / "config.json").write_text(json.dumps(config))
    pocs = run_company_poc_job(
        job_dir,
        source_builder=lambda _names: [SuccessfulGoogleSource()],
    )
    status = read_status(job_dir)
    assert len(pocs) == 1
    assert status["failed_searches"] == 0


def test_company_poc_migrates_old_error_only_checkpoint_for_retry(tmp_path: Path):
    job_dir = create_job(
        "company_pocs",
        {
            "companies": ["MUFG Bank Singapore"],
            "designations": ["Senior Director"],
            "sources": ["google_browser"],
            "retry_attempts": 1,
            "retry_failed_searches": True,
        },
        jobs_root=tmp_path,
    )
    (job_dir / "checkpoint.json").write_text(
        json.dumps(
            {
                "version": 1,
                "task_index": 1,
                "source_index": 0,
                "pocs": [],
                "rejections": [],
                "errors": ["google_browser: old challenge"],
            }
        )
    )
    pocs = run_company_poc_job(
        job_dir,
        source_builder=lambda _names: [SuccessfulGoogleSource()],
    )
    assert len(pocs) == 1


def test_job_without_a_live_pid_becomes_recoverable():
    assert job_is_stale(
        {
            "state": "running",
            "pid": 0,
            "updated_at": "2020-01-01T00:00:00+00:00",
        }
    )


def test_lead_worker_marks_preflight_errors_failed_instead_of_stuck_running(tmp_path: Path):
    job_dir = create_job(
        "lead_harvest",
        {
            "roles": ["CTO"],
            "locations": ["Bengaluru"],
            "sources": ["google_browser"],
            "require_target_company": True,
            "company_names": [],
        },
        jobs_root=tmp_path,
    )

    with pytest.raises(ValueError, match="Hard company filtering"):
        run_lead_job(job_dir)

    status = read_status(job_dir)
    assert status["state"] == "failed"
    assert status["phase"] == "startup"
    assert "Worker initialization failed" in status["message"]


class StopAfterLeadSearch:
    name = "fake"

    def __init__(self, job_dir: Path, *, stop: bool):
        self.job_dir = job_dir
        self.stop = stop
        self.calls = 0

    def search(self, query, *, max_results, headless=True):
        self.calls += 1
        if self.stop:
            (self.job_dir / "stop.requested").touch()
        return [
            SearchResult(
                title="Asha Rao - Chief Technology Officer - Razorpay | LinkedIn",
                body="Location: Bengaluru · Razorpay payments platform.",
                href="https://www.linkedin.com/in/asha-rao/",
                source="fake",
                query=query,
            )
        ]


class LeadChallengeSource:
    name = "google_browser"

    def search(self, query, *, max_results, headless=True):
        raise SourceError("google challenge", disable_source=True, challenge=True)


class LeadRecoverySource:
    name = "google_browser"

    def search(self, query, *, max_results, headless=True):
        return [
            SearchResult(
                title="Asha Rao - Chief Technology Officer - Razorpay | LinkedIn",
                body="Location: Bengaluru · Razorpay payments platform.",
                href="https://www.linkedin.com/in/asha-rao/",
                source=self.name,
                query=query,
            )
        ]


class LeadBingFallbackSource(LeadRecoverySource):
    name = "bing_browser"


def test_lead_google_challenge_adds_bing_fallback_and_continues(tmp_path: Path):
    job_dir = create_job(
        "lead_harvest",
        {
            "roles": ["CTO"],
            "locations": ["Bengaluru"],
            "company_names": ["Razorpay"],
            "require_target_company": True,
            "sources": ["google_browser"],
            "target_count": 1,
            "max_queries": 1,
        },
        jobs_root=tmp_path,
    )

    def source_builder(names):
        return [LeadChallengeSource()] if names == ["google_browser"] else [LeadBingFallbackSource()]

    result = run_lead_job(job_dir, source_builder=source_builder)

    status = read_status(job_dir)
    config = json.loads((job_dir / "config.json").read_text())
    assert [lead.name for lead in result.leads] == ["Asha Rao"]
    assert status["state"] == "completed_with_warnings"
    assert "bing_browser" in config["sources"]
    assert status["failed_searches"] == 1


def test_lead_warning_recovery_replays_completed_checkpoint(tmp_path: Path):
    job_dir = create_job(
        "lead_harvest",
        {
            "roles": ["CTO"],
            "locations": ["Bengaluru"],
            "company_names": ["Razorpay"],
            "require_target_company": True,
            "sources": ["google_browser"],
            "target_count": 1,
            "max_queries": 1,
        },
        jobs_root=tmp_path,
    )

    run_lead_job(job_dir, source_builder=lambda _names: [LeadChallengeSource()])
    first_status = read_status(job_dir)
    assert first_status["state"] == "completed_with_warnings"
    assert first_status["failed_searches"] == 1

    config = json.loads((job_dir / "config.json").read_text())
    config["retry_failed_searches"] = True
    config["browser_headless"] = False
    config["google_manual_challenge_seconds"] = 1
    (job_dir / "config.json").write_text(json.dumps(config))

    result = run_lead_job(job_dir, source_builder=lambda _names: [LeadRecoverySource()])

    final_status = read_status(job_dir)
    assert [lead.name for lead in result.leads] == ["Asha Rao"]
    assert final_status["state"] == "completed"
    assert final_status["failed_searches"] == 0


def test_main_lead_job_checkpoints_search_and_resumes_into_verification(tmp_path: Path):
    dedupe_path = tmp_path / "existing.csv"
    dedupe_path.write_text(
        "Name,LinkedIn URL\n"
        "Existing Person,https://www.linkedin.com/in/existing-person/\n"
        "Malformed Value,https://[2001:db8::1\n"
        "IP Value,192.0.2.10\n",
        encoding="utf-8",
    )
    job_dir = create_job(
        "lead_harvest",
        {
            "roles": ["CTO"],
            "locations": ["Bengaluru"],
            "industries": [],
            "company_names": ["Razorpay"],
            "require_target_company": True,
            "minimum_confidence": 80,
            "sources": ["fake"],
            "target_count": 1,
            "max_queries": 1,
            "existing_files": [str(dedupe_path)],
        },
        jobs_root=tmp_path,
    )
    first = StopAfterLeadSearch(job_dir, stop=True)

    run_lead_job(job_dir, source_builder=lambda _names: [first])

    assert read_status(job_dir)["state"] == "paused"
    result, checkpoint = load_lead_job_checkpoint(job_dir)
    assert checkpoint["phase"] == "search"
    assert checkpoint["query_index"] == 1
    assert result.leads == []

    clear_stop(job_dir)
    second = StopAfterLeadSearch(job_dir, stop=False)
    result = run_lead_job(job_dir, source_builder=lambda _names: [second])

    assert second.calls == 0
    final_status = read_status(job_dir)
    assert final_status["state"] == "completed"
    assert [lead.name for lead in result.leads] == ["Asha Rao"]
    workbook = load_workbook(final_status["xlsx_path"], read_only=True)
    assert "Filter Contract" in workbook.sheetnames


class PagedLeadSource:
    name = "paged"

    def __init__(self, job_dir: Path, *, stop_on_first_page: bool):
        self.job_dir = job_dir
        self.stop_on_first_page = stop_on_first_page
        self.pages = []

    def search_page(self, query, *, page, max_results, headless=True):
        self.pages.append(page)
        if page == 1:
            name, company, slug = "Asha Rao", "Razorpay", "asha-rao"
            if self.stop_on_first_page:
                (self.job_dir / "stop.requested").touch()
        else:
            name, company, slug = "Bina Shah", "PhonePe", "bina-shah"
        return SearchPage(
            results=[
                SearchResult(
                    title=f"{name} - Chief Technology Officer - {company} | LinkedIn",
                    body=f"Location: Bengaluru · {company} payments platform.",
                    href=f"https://www.linkedin.com/in/{slug}/",
                    source=self.name,
                    query=query,
                )
            ],
            page=page,
            has_next=page == 1,
        )


def test_main_lead_job_resumes_at_exact_result_page(tmp_path: Path):
    job_dir = create_job(
        "lead_harvest",
        {
            "roles": ["CTO"],
            "locations": ["Bengaluru"],
            "industries": [],
            "company_names": [],
            "require_target_company": False,
            "minimum_confidence": 0,
            "sources": ["paged"],
            "target_count": 2,
            "max_queries": 1,
            "max_results_per_query": 20,
            "max_pages_per_query": 2,
        },
        jobs_root=tmp_path,
    )
    first = PagedLeadSource(job_dir, stop_on_first_page=True)

    run_lead_job(job_dir, source_builder=lambda _names: [first])

    _result, checkpoint = load_lead_job_checkpoint(job_dir)
    assert read_status(job_dir)["state"] == "paused"
    assert first.pages == [1]
    assert checkpoint["query_index"] == 0
    assert checkpoint["page_index"] == 1

    clear_stop(job_dir)
    second = PagedLeadSource(job_dir, stop_on_first_page=False)
    result = run_lead_job(job_dir, source_builder=lambda _names: [second])

    assert second.pages == [2]
    assert {lead.name for lead in result.leads} == {"Asha Rao", "Bina Shah"}
