from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from speedy_scraper.background_jobs import create_job, read_status, write_json
from speedy_scraper.outreach_intelligence import (
    MatchType,
    OutreachDecision,
    ReviewAction,
    SourceRole,
    SourceSpec,
    StatusCategory,
    apply_review_overrides,
    build_outreach_frames,
    canonicalize_frame,
    detect_columns,
    load_review_decisions,
    normalize_company,
    normalize_email,
    normalize_identity_text,
    normalize_phone,
    parse_google_sheet_url,
    read_source_bytes,
    run_outreach_match,
    save_review_decisions,
    suggest_status_category,
    write_outreach_exports,
)
from speedy_scraper.outreach_job import load_outreach_checkpoint, run_outreach_job


def _canonical_records(
    rows: list[dict[str, object]],
    *,
    role: SourceRole,
    source_id: str,
    mapping: dict[str, str] | None = None,
):
    frame = pd.DataFrame(rows)
    selected_mapping = mapping or detect_columns(frame.columns).mapping
    spec = SourceSpec(
        source_id=source_id,
        source_name=f"{source_id}.csv",
        path="",
        sheet_name="CSV",
        role=role,
        mapping=selected_mapping,
    )
    return canonicalize_frame(frame, spec)


def test_column_detection_maps_name_and_poc_to_the_same_canonical_field():
    primary = detect_columns(["Name", "Company", "Work Email"])
    previous = detect_columns(["POC", "Organisation", "POC Status"])

    assert primary.mapping["full_name"] == "Name"
    assert previous.mapping["full_name"] == "POC"
    assert primary.mapping["email"] == "Work Email"
    assert previous.mapping["status"] == "POC Status"


def test_column_detection_is_case_and_punctuation_insensitive_and_flags_ambiguity():
    detection = detect_columns(["FULL_NAME", "Poc", "linkedin-url"])

    assert detection.mapping["linkedin_url"] == "linkedin-url"
    assert "full_name" in detection.ambiguous
    assert set(detection.ambiguous["full_name"]) == {"FULL_NAME", "Poc"}


def test_first_and_last_name_reconstruction_preserves_unicode():
    records, invalid = _canonical_records(
        [{"Given Name": "José", "Surname": "D'Souza", "Firm": "Ácme Pvt Ltd"}],
        role=SourceRole.PRIMARY,
        source_id="primary",
    )

    assert invalid == []
    assert records[0].full_name == "José D'Souza"
    assert records[0].normalized_name == "jose d souza"
    assert records[0].normalized_company == "acme"


def test_phone_email_company_and_status_normalization():
    assert normalize_phone("+91 98765 43210") == normalize_phone("9876543210", "IN")
    assert normalize_email(" Jane@Example.COM ") == "jane@example.com"
    assert normalize_email("not-an-email") == ""
    assert normalize_company("Example Technologies Private Limited") == "example technologies"
    assert suggest_status_category("Do not contact") == StatusCategory.DO_NOT_CONTACT
    assert suggest_status_category("Interested") == StatusCategory.POSITIVE_RESPONSE


@pytest.mark.parametrize(
    ("field", "value", "expected_type", "confidence"),
    [
        ("Phone", "+91 98765 43210", MatchType.EXACT_PHONE, 100),
        ("Email", "jane@example.com", MatchType.EXACT_EMAIL, 99),
        (
            "LinkedIn URL",
            "https://www.linkedin.com/in/jane-doe/?trk=test",
            MatchType.EXACT_LINKEDIN,
            98,
        ),
    ],
)
def test_exact_identifiers_use_the_required_priority(field, value, expected_type, confidence):
    primary_rows = [{"Name": "Jane Doe", "Company": "Example", field: value}]
    previous_value = value
    if field == "Phone":
        previous_value = "9876543210"
    if field == "LinkedIn URL":
        previous_value = "linkedin.com/in/jane-doe/"
    previous_rows = [
        {"POC": "Jane Doe", "Company": "Example", field: previous_value, "Status": "Contacted"}
    ]
    primary, _ = _canonical_records(
        primary_rows, role=SourceRole.PRIMARY, source_id="primary"
    )
    previous, _ = _canonical_records(
        previous_rows, role=SourceRole.PREVIOUS, source_id="previous"
    )

    run = run_outreach_match(primary, previous)
    result = run.matches[0]

    assert result.match_type == expected_type
    assert result.confidence == confidence
    assert result.confirmed is True
    assert result.decision == OutreachDecision.ALREADY_CONTACTED


def test_exact_name_company_matches_after_legal_suffix_normalization():
    primary, _ = _canonical_records(
        [{"Name": "Asha Rao", "Company": "Example Pvt Ltd"}],
        role=SourceRole.PRIMARY,
        source_id="primary",
    )
    previous, _ = _canonical_records(
        [{"POC": "Asha Rao", "Organisation": "Example Private Limited", "Status": "Contacted"}],
        role=SourceRole.PREVIOUS,
        source_id="previous",
    )

    result = run_outreach_match(primary, previous).matches[0]

    assert result.match_type == MatchType.NAME_COMPANY_MATCH
    assert result.confidence == 90
    assert result.confirmed is True


def test_fuzzy_name_company_is_review_only_and_different_company_is_not_matched():
    primary, _ = _canonical_records(
        [
            {"Name": "Jonathon Smith", "Company": "Razorpay"},
            {"Name": "Jonathon Smith", "Company": "Unrelated Holdings"},
        ],
        role=SourceRole.PRIMARY,
        source_id="primary",
    )
    previous, _ = _canonical_records(
        [{"POC": "Jonathan Smith", "Company": "Razorpay Technologies", "Status": "Contacted"}],
        role=SourceRole.PREVIOUS,
        source_id="previous",
    )

    run = run_outreach_match(primary, previous)

    assert run.matches[0].match_type == MatchType.POSSIBLE_MATCH
    assert run.matches[0].decision == OutreachDecision.POSSIBLE_MATCH
    assert run.matches[0].confirmed is False
    assert run.matches[1].match_type == MatchType.NO_MATCH
    assert run.matches[1].decision == OutreachDecision.SAFE_TO_CONTACT


def test_name_only_fuzzy_match_never_auto_excludes():
    primary, _ = _canonical_records(
        [{"Name": "Sanjay Mehta"}], role=SourceRole.PRIMARY, source_id="primary"
    )
    previous, _ = _canonical_records(
        [{"POC": "Sanjay Mehta", "Status": "Contacted"}],
        role=SourceRole.PREVIOUS,
        source_id="previous",
    )

    result = run_outreach_match(primary, previous).matches[0]

    assert result.match_type == MatchType.POSSIBLE_MATCH
    assert result.decision == OutreachDecision.POSSIBLE_MATCH
    assert result.confirmed is False


def test_conflicting_strong_identifiers_are_routed_to_review():
    primary, _ = _canonical_records(
        [{"Name": "Asha Rao", "Email": "asha@example.com", "Phone": "9876543210"}],
        role=SourceRole.PRIMARY,
        source_id="primary",
    )
    previous, _ = _canonical_records(
        [
            {"POC": "Asha Rao", "Email": "asha@example.com", "Phone": "9123456789"},
            {"POC": "Another Person", "Email": "other@example.com", "Phone": "9876543210"},
        ],
        role=SourceRole.PREVIOUS,
        source_id="previous",
    )

    result = run_outreach_match(primary, previous).matches[0]

    assert result.decision == OutreachDecision.REVIEW_REQUIRED
    assert result.confirmed is False
    assert len(result.previous_ids) == 2
    frames = build_outreach_frames(run_outreach_match(primary, previous))
    rejected = frames["invalid"]
    assert "conflicting_or_ambiguous_identity" in set(rejected["Reason"])
    assert "Asha Rao" in frames["review"].iloc[0]["Previous Name"]
    assert "Another Person" in frames["review"].iloc[0]["Previous Name"]


def test_exact_identifier_with_materially_different_name_requires_review():
    primary, _ = _canonical_records(
        [{"Name": "Jane Doe", "Email": "shared@example.com"}],
        role=SourceRole.PRIMARY,
        source_id="primary",
    )
    previous, _ = _canonical_records(
        [{"POC": "Completely Different", "Email": "shared@example.com", "Status": "Contacted"}],
        role=SourceRole.PREVIOUS,
        source_id="previous",
    )

    result = run_outreach_match(primary, previous).matches[0]

    assert result.match_type == MatchType.EXACT_EMAIL
    assert result.decision == OutreachDecision.REVIEW_REQUIRED
    assert result.confirmed is False


def test_multiple_previous_interactions_are_preserved():
    primary, _ = _canonical_records(
        [{"Name": "Jane Doe", "Email": "jane@example.com"}],
        role=SourceRole.PRIMARY,
        source_id="primary",
    )
    previous_a, _ = _canonical_records(
        [{"POC": "Jane Doe", "Email": "jane@example.com", "Status": "Contacted", "Notes": "Event A"}],
        role=SourceRole.PREVIOUS,
        source_id="event_a",
    )
    previous_b, _ = _canonical_records(
        [{"POC": "Jane Doe", "Email": "jane@example.com", "Status": "Interested", "Notes": "Event B"}],
        role=SourceRole.PREVIOUS,
        source_id="event_b",
    )

    run = run_outreach_match(primary, [*previous_a, *previous_b])
    frames = build_outreach_frames(run)

    assert len(run.matches[0].previous_ids) == 2
    assert len(frames["common"]) == 2
    assert set(frames["common"]["Previous Notes"]) == {"Event A", "Event B"}


def test_primary_duplicates_are_skipped_but_previous_duplicates_remain_history():
    primary, _ = _canonical_records(
        [
            {"Name": "Jane Doe", "Email": "jane@example.com"},
            {"Name": "Jane Doe", "Email": "jane@example.com"},
        ],
        role=SourceRole.PRIMARY,
        source_id="primary",
    )
    previous, _ = _canonical_records(
        [
            {"POC": "Jane Doe", "Email": "jane@example.com", "Status": "Contacted"},
            {"POC": "Jane Doe", "Email": "jane@example.com", "Status": "No response"},
        ],
        role=SourceRole.PREVIOUS,
        source_id="previous",
    )

    run = run_outreach_match(primary, previous)

    assert len(run.primary_records) == 1
    assert len(run.previous_records) == 2
    assert run.metrics["duplicates"] == 2
    assert {row.reason for row in run.invalid_rows} >= {
        "duplicate_primary_record",
        "duplicate_previous_record",
    }


def test_blank_status_defaults_to_review_and_custom_mapping_can_make_it_safe():
    primary, _ = _canonical_records(
        [{"Name": "Jane Doe", "Email": "jane@example.com"}],
        role=SourceRole.PRIMARY,
        source_id="primary",
    )
    previous, _ = _canonical_records(
        [{"POC": "Jane Doe", "Email": "jane@example.com", "Status": "Can invite again"}],
        role=SourceRole.PREVIOUS,
        source_id="previous",
    )

    default_run = run_outreach_match(primary, previous)
    safe_run = run_outreach_match(
        primary,
        previous,
        status_map={"Can invite again": StatusCategory.SAFE_OR_RECONTACT_ALLOWED.value},
    )

    assert default_run.matches[0].decision == OutreachDecision.REVIEW_REQUIRED
    assert safe_run.matches[0].decision == OutreachDecision.SAFE_TO_CONTACT


def test_malformed_identifiers_are_reported_without_losing_a_named_row():
    records, invalid = _canonical_records(
        [{"Name": "Jane Doe", "Email": "broken", "LinkedIn": "https://example.com/jane"}],
        role=SourceRole.PRIMARY,
        source_id="primary",
    )

    assert len(records) == 1
    assert {row.reason for row in invalid} == {"malformed_email", "invalid_linkedin_url"}
    assert all(row.disposition == "PROCESSED_WITH_WARNING" for row in invalid)


def test_csv_and_multisheet_xlsx_loading():
    csv_sheets = read_source_bytes(b"POC,Company\nJane,Example\n", "people.csv")
    buffer = Path(__file__).parent / "_not_written.xlsx"
    del buffer
    from io import BytesIO

    workbook = BytesIO()
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        pd.DataFrame({"Name": ["A"]}).to_excel(writer, sheet_name="Primary", index=False)
        pd.DataFrame({"POC": ["B"]}).to_excel(writer, sheet_name="History", index=False)
    xlsx_sheets = read_source_bytes(workbook.getvalue(), "people.xlsx")

    assert list(csv_sheets) == ["CSV"]
    assert set(xlsx_sheets) == {"Primary", "History"}


def test_google_sheet_url_parses_gid_and_rejects_non_google_hosts():
    sheet_id, gid, export_url = parse_google_sheet_url(
        "https://docs.google.com/spreadsheets/d/abc_123/edit#gid=456"
    )

    assert sheet_id == "abc_123"
    assert gid == "456"
    assert export_url.endswith("format=csv&gid=456")
    with pytest.raises(ValueError):
        parse_google_sheet_url("https://example.com/spreadsheets/d/abc/edit#gid=0")


def test_review_overrides_change_decisions_and_persist_atomically(tmp_path: Path):
    primary, _ = _canonical_records(
        [{"Name": "Sanjay Mehta"}], role=SourceRole.PRIMARY, source_id="primary"
    )
    previous, _ = _canonical_records(
        [{"POC": "Sanjay Mehta", "Status": "Contacted"}],
        role=SourceRole.PREVIOUS,
        source_id="previous",
    )
    run = run_outreach_match(primary, previous)
    record_id = run.matches[0].primary_id
    overrides = {record_id: ReviewAction.REJECT_MATCH.value}

    save_review_decisions(tmp_path, overrides)
    revised = apply_review_overrides(run, load_review_decisions(tmp_path))

    assert revised.matches[0].decision == OutreachDecision.SAFE_TO_CONTACT
    assert revised.matches[0].confirmed is False


def test_exports_include_fresh_common_combined_invalid_history_and_config(tmp_path: Path):
    primary, invalid = _canonical_records(
        [
            {"Name": "Jane Doe", "Email": "jane@example.com"},
            {"Name": "New Person", "Email": "new@example.com"},
            {"Name": "Broken", "Email": "bad"},
        ],
        role=SourceRole.PRIMARY,
        source_id="primary",
    )
    previous, previous_invalid = _canonical_records(
        [{"POC": "Jane Doe", "Email": "jane@example.com", "Status": "Contacted"}],
        role=SourceRole.PREVIOUS,
        source_id="previous",
    )
    run = run_outreach_match(
        primary, previous, invalid_rows=[*invalid, *previous_invalid], config={"region": "IN"}
    )

    outputs = write_outreach_exports(run, tmp_path)
    workbook = load_workbook(outputs["xlsx"], read_only=True)
    fresh = pd.read_csv(outputs["fresh_csv"])
    combined = pd.read_csv(outputs["combined_csv"])

    assert set(workbook.sheetnames) == {
        "Fresh Outreach",
        "Common People",
        "Review Queue",
        "Combined Master",
        "Invalid Rows",
        "Interaction History",
        "Run Configuration",
    }
    assert set(fresh["Name"]) == {"New Person", "Broken"}
    assert set(combined["Membership"]) == {"BOTH", "PRIMARY_ONLY"}
    assert Path(outputs["invalid_csv"]).exists()


def test_combined_master_includes_previous_only_people():
    primary, _ = _canonical_records(
        [{"Name": "Jane Doe", "Email": "jane@example.com"}],
        role=SourceRole.PRIMARY,
        source_id="primary",
    )
    previous, _ = _canonical_records(
        [
            {"POC": "Jane Doe", "Email": "jane@example.com", "Status": "Contacted"},
            {"POC": "Old Person", "Email": "old@example.com", "Status": "Contacted"},
        ],
        role=SourceRole.PREVIOUS,
        source_id="previous",
    )

    combined = build_outreach_frames(run_outreach_match(primary, previous))["combined"]

    assert set(combined["Membership"]) == {"BOTH", "PREVIOUS_ONLY"}


def test_background_job_writes_checkpoint_status_and_exports(tmp_path: Path):
    primary_path = tmp_path / "primary.csv"
    previous_path = tmp_path / "previous.csv"
    pd.DataFrame(
        {"Name": ["Jane Doe", "New Person"], "Email": ["jane@example.com", "new@example.com"]}
    ).to_csv(primary_path, index=False)
    pd.DataFrame(
        {"POC": ["Jane Doe"], "Email": ["jane@example.com"], "Status": ["Contacted"]}
    ).to_csv(previous_path, index=False)
    job_dir = create_job("outreach_intelligence", {}, jobs_root=tmp_path / "jobs")
    write_json(
        job_dir / "config.json",
        {
            "default_phone_region": "IN",
            "status_map": {"Contacted": StatusCategory.CONTACTED.value},
            "primary": {
                "source_id": "primary",
                "source_name": primary_path.name,
                "path": str(primary_path),
                "sheet_name": "CSV",
                "role": SourceRole.PRIMARY.value,
                "mapping": {"full_name": "Name", "email": "Email"},
            },
            "previous": [
                {
                    "source_id": "previous",
                    "source_name": previous_path.name,
                    "path": str(previous_path),
                    "sheet_name": "CSV",
                    "role": SourceRole.PREVIOUS.value,
                    "mapping": {
                        "full_name": "POC",
                        "email": "Email",
                        "status": "Status",
                    },
                }
            ],
        },
    )

    run = run_outreach_job(job_dir)
    loaded, checkpoint = load_outreach_checkpoint(job_dir)
    status = read_status(job_dir)

    assert run.metrics["safe_to_contact"] == 1
    assert loaded is not None
    assert checkpoint["version"] == 1
    assert status["state"] == "completed"
    assert Path(status["outputs"]["xlsx"]).exists()


def test_normalized_text_keeps_non_latin_names_usable():
    assert normalize_identity_text("राम कुमार")
